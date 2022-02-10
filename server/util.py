# -*- coding: utf8 -*-
#from werkzeug._compat import text_type, PY2
import os
import re
from openpyxl import Workbook,styles
import numpy as np
_filename_ascii_strip_re = re.compile(r"[^A-Za-z0-9_.-]")
_windows_device_files = (
    "CON",
    "AUX",
    "COM1",
    "COM2",
    "COM3",
    "COM4",
    "LPT1",
    "LPT2",
    "LPT3",
    "PRN",
    "NUL",
)


def secure_filename(filename):
    r"""Pass it a filename and it will return a secure version of it.  This
    filename can then safely be stored on a regular file system and passed
    to :func:`os.path.join`.  The filename returned is an ASCII only string
    for maximum portability.

    On windows systems the function also makes sure that the file is not
    named after one of the special device files.

    >>> secure_filename("My cool movie.mov")
    'My_cool_movie.mov'
    >>> secure_filename("../../../etc/passwd")
    'etc_passwd'
    >>> secure_filename(u'i contain cool \xfcml\xe4uts.txt')
    'i_contain_cool_umlauts.txt'

    The function might return an empty filename.  It's your responsibility
    to ensure that the filename is unique and that you generate random
    filename if the function returned an empty one.

    .. versionadded:: 0.5

    :param filename: the filename to secure
    """
    #if isinstance(filename, text_type):
    from unicodedata import normalize
    filename = normalize('NFKD', filename).encode('utf-8', 'ignore')  # 转码
    #    if not PY2:
    filename = filename.decode('utf-8')
    for sep in os.path.sep, os.path.altsep:
        if sep:
            filename = filename.replace(sep, ' ')
    _filename_ascii_add_strip_re = re.compile(r'[^A-Za-z0-9_\u4E00-\u9FBF.-]')
    filename = str(_filename_ascii_add_strip_re.sub('', '_'.join(
        filename.split()))).strip('._')

    return filename

def point_in_range(point,range_list):
    for rec in range_list:
        if (point[0]>=rec[0] and point[0]<=rec[2]) and (point[1]>=rec[1] and point[1]<=rec[3]):
            return True,rec
    return False,None

def dict2xls(tabinfo_list,xlspath):
    wb = Workbook()
    wb.remove(wb.active)
    for ti,tabinfo in enumerate(tabinfo_list):
        ws = wb.create_sheet(title='Tab_'+str(ti))
        colnum = 0
        max_col_rkey = ''
        row_num = 0

        for rkey in tabinfo:
            if colnum<len(tabinfo[rkey]):
                colnum = len(tabinfo[rkey])
                max_col_rkey = rkey
            row_num += 1
        max_col_pos = []

        for colinfo in tabinfo[max_col_rkey]:
            max_col_pos.append(colinfo['cell_pos'])

        #获得列细分pos和行细分pos
        base_row_pos = []
        #row_diff = []
        for rkey in tabinfo:
            row_h = []
            matcher = {}
            for i,colinfo in enumerate(tabinfo[rkey]):
                row_h.append((colinfo['cell_pos'][5]-colinfo['cell_pos'][3]+colinfo['cell_pos'][7]-colinfo['cell_pos'][1])/2)
                index = np.argmin(np.abs(colinfo['cell_pos'][0]-np.array(max_col_pos)[:,0]))
                if index not in matcher:
                    matcher[index] = i
                else:
                    max_col_pos.pop(index)
                    max_col_pos.insert(index, tabinfo[rkey][matcher[index]]['cell_pos'])
                    max_col_pos.insert(index + 1, colinfo['cell_pos'])
                    matcher[index + 1] = i
            base_row_pos.append(tabinfo[rkey][np.argmin(row_h)]['cell_pos'])
            #row_diff.append(max(row_h)>1.8*min(row_h))

        max_col_pos = np.array(max_col_pos)
        merged_cell_ranges = []
        for rkey in tabinfo:
            ri = int(rkey.split('_')[1])
            for col_i,colinfo in enumerate(tabinfo[rkey]):
                # if len(tabinfo[rkey])==colnum and not row_diff[ri-1]:
                #     ci = colinfo['col_index']+1
                #     if colinfo['text'] != '' and colinfo['text'] != ' ':
                #         try:
                #             ws.cell(row=ri,column=ci,value=colinfo['text'])
                #         except:
                #             if len(merged_cell_ranges)>0:
                #                 flag,merge_range = point_in_range([ri,ci],merged_cell_ranges)
                #                 if flag:
                #                     ws.unmerge_cells(start_row=merge_range[0], start_column=merge_range[1], end_row=merge_range[2], end_column=merge_range[3])
                #                     merged_cell_ranges.remove(merge_range)
                #                     ws.cell(row=ri, column=ci, value=colinfo['text'])
                
                if int(max_col_rkey.split('_')[1])>ri:
                    ci = np.argmin(np.abs(colinfo['cell_pos'][6]-max_col_pos[:,0]))+1
                    ci_e = np.argmin(np.abs(colinfo['cell_pos'][4]-max_col_pos[:,2]))+1
                    while ci_e<ci:
                        ci_e += 1
                else:
                    ci = np.argmin(np.abs(colinfo['cell_pos'][0]-max_col_pos[:,6]))+1
                    ci_e = np.argmin(np.abs(colinfo['cell_pos'][2] - max_col_pos[:, 4])) + 1
                    while ci_e<ci:
                        ci_e += 1
                if colinfo['text']!='' and colinfo['text']!=' ':
                    try:
                        ws.cell(row=ri, column=ci, value=colinfo['text'])
                    except:
                        if len(merged_cell_ranges) > 0:
                            flag, merge_range = point_in_range([ri, ci], merged_cell_ranges)
                            if flag:
                                ws.unmerge_cells(start_row=merge_range[0], start_column=merge_range[1],
                                                    end_row=merge_range[2], end_column=merge_range[3])
                                merged_cell_ranges.remove(merge_range)
                                ws.cell(row=ri, column=ci, value=colinfo['text'])
                start_row,start_column,end_row,end_column = ri, ci, ri, ci
                if ci_e!=ci:
                    end_column = ci_e
                if len(tabinfo[rkey])>1:
                    end_row = np.argmin(np.abs(colinfo['cell_pos'][5] - np.array(base_row_pos)[:, 5])) + 1
                if (start_row!=end_row or start_column!=end_column) and (start_row<=end_row and start_column<=end_column):
                    #print(start_row,start_column,end_row,end_column)
                    ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
                    merged_cell_ranges.append([start_row,start_column,end_row,end_column])
        for rr in ws:
            for cc in rr:
                cc.alignment = styles.Alignment(horizontal='center', vertical='center',wrapText=True)
    wb.save(xlspath)
